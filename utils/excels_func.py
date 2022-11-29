from openpyxl import load_workbook, Workbook


def create_excel(name, sheet_name, data):
    try:
        wb = load_workbook(f'{name}.xlsx')
        try:
            ws = wb.get_sheet_by_name(sheet_name)
            wb.remove(ws)
            wb.create_sheet(sheet_name)
            ws = wb.get_sheet_by_name(sheet_name)
        except:
            wb.create_sheet(sheet_name)
            ws = wb.get_sheet_by_name(sheet_name)
            print('created new sheet')
    except:
        wb = Workbook()
        if wb.active.title == 'Sheet':
            wb.remove(wb.active)
        wb.create_sheet(sheet_name)
        ws = wb.get_sheet_by_name(sheet_name)
        print('created new excel')

    for i in data:
        ws.append(i)
    wb.save(f'{name}.xlsx')
